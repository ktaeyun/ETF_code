"""
analysis/pairwise_regime_cooccurrence.py
=========================================
5개 외생변수 HMM 레짐 레이블 간 공통빈도 분석

변수:
  - Global_RV_regime       : 일별 (Global_RV HMM Viterbi)
  - VKOSPI_resid_regime    : 일별 (VKOSPI HAR 잔차 HMM Viterbi, lag=22 offset)
  - domestic_btc_svi_regime: 주별 → 일별 forward-fill
  - global_btc_svi_regime  : 주별 → 일별 forward-fill
  - btc_volume_btc_regime  : 주별 → 일별 forward-fill

출력: results/pairwise_regime_cooccurrence_results.xlsx
"""

from __future__ import annotations

import itertools
import sys
from pathlib import Path

import numpy as np
import pandas as pd

_ROOT = Path(__file__).resolve().parent.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from preprocessing.run_pipeline import load_data, step_har, step_hmm_vol, step_hmm_svi


# ══════════════════════════════════════════════════════════════
# 상수
# ══════════════════════════════════════════════════════════════

REGIME_COLS = [
    "Global_RV_regime",
    "VKOSPI_resid_regime",
    "domestic_btc_svi_regime",
    "global_btc_svi_regime",
    "btc_volume_btc_regime",
]

SAVE_PATH = _ROOT / "results" / "pairwise_regime_cooccurrence_results.xlsx"

# 시나리오 외생변수 생성용
_COL_MAP: dict[str, str] = {
    "Global_RV_regime":        "Bitcoin_RV",
    "VKOSPI_resid_regime":     "VKOSPI",
    "btc_volume_btc_regime":   "KR_Volume",
    "domestic_btc_svi_regime": "KR_SVI",
    "global_btc_svi_regime":   "Global_SVI",
}
_DISPLAY_COLS = ["Bitcoin_RV", "VKOSPI", "KR_Volume", "KR_SVI", "Global_SVI"]

SCENARIO_CSV_PATH = _ROOT / "results" / "scenario_selection" / "final_scenarios_latest.csv"
EXOG_SAVE_PATH    = _ROOT / "results" / "scenario_exog_vars"


# ══════════════════════════════════════════════════════════════
# Step 1: HMM 레짐 DataFrame 구축
# ══════════════════════════════════════════════════════════════

def _optimal_states(comp) -> np.ndarray:
    """HMMComparison → 최적 K의 Viterbi 상태 배열."""
    return {
        1: comp.result_k1,
        2: comp.result_k2,
        3: comp.result_k3,
    }[comp.optimal_K].states


def build_regime_df(
    n_init: int = 10,
    B: int = 1000,
) -> pd.DataFrame:
    """5개 변수 HMM 레짐 레이블을 일별 DatetimeIndex로 정렬한 DataFrame 반환.

    주별(W-MON) SVI/Volume 레짐은 forward-fill 로 일별 확장.

    Returns
    -------
    pd.DataFrame : DatetimeIndex, 컬럼 = REGIME_COLS (Int64, NaN 가능)
    """
    print("=" * 60)
    print("  [Step 1] 데이터 로드")
    print("=" * 60)
    df_daily, df_weekly = load_data()

    print("\n" + "=" * 60)
    print("  [Step 2] HAR-VKOSPI 잔차 추출")
    print("=" * 60)
    har_result = step_har(df_daily, save=False)

    print("\n" + "=" * 60)
    print("  [Step 3] HMM — SVI/Volume (주별)")
    print("=" * 60)
    hmm_svi = step_hmm_svi(df_weekly, n_init=n_init, B=B, save=False)

    print("\n" + "=" * 60)
    print("  [Step 4] HMM — 변동성 (일별)")
    print("=" * 60)
    hmm_vol = step_hmm_vol(df_daily, har_result, n_init=n_init, B=B, save=False)

    # 일별 인덱스
    daily_idx = df_daily.index   # df_daily는 dropna() 완료

    # ── Global_RV regime (일별) ────────────────────────────────
    states_rv = _optimal_states(hmm_vol["Global_RV"])
    global_rv_regime = pd.Series(
        states_rv,
        index=daily_idx[: len(states_rv)],
        name="Global_RV_regime",
        dtype="Int64",
    )

    # ── VKOSPI_resid regime (일별, HAR lag offset=22) ──────────
    states_vk = _optimal_states(hmm_vol["VKOSPI_resid"])
    n_vk = len(states_vk)
    vkospi_regime = pd.Series(
        states_vk,
        index=daily_idx[22 : 22 + n_vk],
        name="VKOSPI_resid_regime",
        dtype="Int64",
    )

    # ── SVI/Volume regimes (주별 → 일별 forward-fill) ───────────
    svi_series: dict[str, pd.Series] = {}
    svi_map = {
        "global_btc_svi"   : "global_btc_svi_regime",
        "domestic_btc_svi" : "domestic_btc_svi_regime",
        "btc_volume_btc"   : "btc_volume_btc_regime",
    }
    for col, col_regime in svi_map.items():
        if col not in hmm_svi:
            print(f"  [!] {col} HMM 결과 없음 — 건너뜀")
            continue
        states = _optimal_states(hmm_svi[col])
        w_idx  = df_weekly.index[: len(states)]
        s_weekly = pd.Series(states, index=w_idx, name=col_regime, dtype="Int64")
        # 주별 → 일별: 공통 날짜에 값 놓고 나머지 forward-fill
        s_daily  = s_weekly.reindex(daily_idx, method="ffill")
        svi_series[col_regime] = s_daily

    # ── 합치기 ─────────────────────────────────────────────────
    all_series = [global_rv_regime, vkospi_regime] + list(svi_series.values())
    df = pd.concat(all_series, axis=1)
    df.index.name = "Date"

    # ── 요약 출력 ──────────────────────────────────────────────
    print("\n" + "=" * 60)
    print("  [Regime DataFrame 요약]")
    print("=" * 60)
    print(f"  shape = {df.shape}")
    print(f"  기간  = {df.index[0].date()} ~ {df.index[-1].date()}")
    print("\n  결측 현황:")
    print(df.isna().sum().rename("NaN 수").to_string())
    print("\n  레짐 분포 (비율):")
    for col in df.columns:
        s = df[col].dropna()
        K = int(s.nunique())
        vc = s.value_counts(normalize=True).sort_index()
        vc_str = "  ".join([f"R{int(r)}={v:.1%}" for r, v in vc.items()])
        print(f"    {col.replace('_regime',''):25s}  K={K}  {vc_str}")

    return df


# ══════════════════════════════════════════════════════════════
# Step 2: 공통빈도 지표 계산
# ══════════════════════════════════════════════════════════════

def _lift_grade(n_AB: int, lift: float) -> str:
    """Lift 값 기준 경험적 실현가능성 보조등급."""
    if n_AB == 0:
        return "공통발생 없음"
    if np.isnan(lift):
        return "N/A"
    if lift >= 1.0:
        return "경험적 근거 강함"
    return "경험적 근거 약함"


def compute_pairwise_cooccurrence(df: pd.DataFrame) -> pd.DataFrame:
    """모든 변수쌍 × 레짐 상태쌍 공통빈도 지표 계산.

    Parameters
    ----------
    df : REGIME_COLS 컬럼을 가진 DatetimeIndex DataFrame

    Returns
    -------
    long-format DataFrame
    """
    vars_ = [c for c in REGIME_COLS if c in df.columns]
    if len(vars_) < 2:
        raise ValueError(f"분석 가능한 변수가 2개 미만입니다: {vars_}")

    rows: list[dict] = []

    for var_A, var_B in itertools.combinations(vars_, 2):
        # 두 변수 모두 유효한 공통 timestep 추출
        pair = df[[var_A, var_B]].dropna()
        T    = len(pair)
        if T == 0:
            print(f"  [!] {var_A} × {var_B}: 공통 관측 없음 — 건너뜀")
            continue

        s_A = pair[var_A].astype(int)
        s_B = pair[var_B].astype(int)

        regimes_A = sorted(s_A.unique())
        regimes_B = sorted(s_B.unique())

        for r_A, r_B in itertools.product(regimes_A, regimes_B):
            mask_A = s_A == r_A
            mask_B = s_B == r_B
            n_A    = int(mask_A.sum())
            n_B    = int(mask_B.sum())
            n_AB   = int((mask_A & mask_B).sum())

            P_A  = n_A  / T
            P_B  = n_B  / T
            P_AB = n_AB / T

            # P_B > 0, P_A > 0 는 unique() 로부터 항상 보장
            P_B_given_A = n_AB / n_A  # = P_AB / P_A
            P_A_given_B = n_AB / n_B  # = P_AB / P_B
            Lift        = P_B_given_A / P_B  # = P_AB / (P_A * P_B)

            rows.append({
                "var_A"          : var_A.replace("_regime", ""),
                "regime_A"       : r_A,
                "var_B"          : var_B.replace("_regime", ""),
                "regime_B"       : r_B,
                "n_A"            : n_A,
                "n_B"            : n_B,
                "n_AB"           : n_AB,
                "T"              : T,
                "P_A"            : round(P_A,  6),
                "P_B"            : round(P_B,  6),
                "P_AB"           : round(P_AB, 6),
                "P_B_given_A"    : round(P_B_given_A, 6),
                "P_A_given_B"    : round(P_A_given_B, 6),
                "Lift"           : round(Lift, 4),
                "empirical_grade": _lift_grade(n_AB, Lift),
            })

    result = pd.DataFrame(rows)
    if not result.empty:
        result = result.sort_values(
            ["var_A", "var_B", "regime_A", "regime_B"]
        ).reset_index(drop=True)

    return result


# ══════════════════════════════════════════════════════════════
# Step 3: Excel 저장 (openpyxl 컬러 스타일 포함)
# ══════════════════════════════════════════════════════════════

_GRADE_COLORS = {
    "경험적 근거 강함": "C6EFCE",   # 녹색
    "경험적 근거 약함": "FFC7CE",   # 빨강
    "공통발생 없음"   : "D9D9D9",   # 회색
}


def _apply_excel_style(ws, df: pd.DataFrame, grade_col: str = "empirical_grade") -> None:
    """openpyxl 워크시트에 헤더 및 등급별 행 색상 적용."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    # 헤더
    hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    # 열 너비 자동 조정
    for col_idx, col_name in enumerate(df.columns, 1):
        max_w = max(
            len(str(col_name)),
            df[col_name].astype(str).str.len().max() if len(df) > 0 else 0,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_w + 2, 28)

    # 등급별 행 색상
    grade_idx = df.columns.get_loc(grade_col) + 1  # 1-based
    for row_i in range(2, len(df) + 2):
        grade = ws.cell(row=row_i, column=grade_idx).value
        color = _GRADE_COLORS.get(grade)
        if color:
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            for col_i in range(1, len(df.columns) + 1):
                ws.cell(row=row_i, column=col_i).fill = fill


def save_to_excel(df_result: pd.DataFrame, df_regime: pd.DataFrame) -> None:
    """분석 결과를 3개 시트 Excel 파일로 저장.

    Sheet 1 '공통빈도 분석' : 전체 long-format 결과
    Sheet 2 '경험적 근거 강함': Lift >= 1.0 행, Lift 내림차순 정렬
    Sheet 3 '레짐 분포'      : 변수별 레짐 분포 요약
    """
    SAVE_PATH.parent.mkdir(parents=True, exist_ok=True)

    # ── Sheet 3용 레짐 분포 요약 ───────────────────────────────
    summary_rows: list[dict] = []
    for col in [c for c in REGIME_COLS if c in df_regime.columns]:
        s = df_regime[col].dropna()
        K = int(s.nunique())
        for r, cnt in s.value_counts().sort_index().items():
            summary_rows.append({
                "variable" : col.replace("_regime", ""),
                "optimal_K": K,
                "regime"   : int(r),
                "n_obs"    : int(cnt),
                "occupancy": round(int(cnt) / len(s), 4),
                "frequency": "주별→일별 forward-fill"
                             if col not in ("Global_RV_regime", "VKOSPI_resid_regime")
                             else "일별",
            })
    df_summary = pd.DataFrame(summary_rows)

    df_strong = (
        df_result[df_result["empirical_grade"] == "경험적 근거 강함"]
        .sort_values("Lift", ascending=False)
        .reset_index(drop=True)
    )

    try:
        with pd.ExcelWriter(str(SAVE_PATH), engine="openpyxl") as writer:
            df_result.to_excel(writer, sheet_name="공통빈도 분석",   index=False)
            df_strong.to_excel(writer, sheet_name="경험적 근거 강함", index=False)
            df_summary.to_excel(writer, sheet_name="레짐 분포",      index=False)

            _apply_excel_style(writer.sheets["공통빈도 분석"],    df_result)
            _apply_excel_style(writer.sheets["경험적 근거 강함"], df_strong)

            # 레짐 분포 시트 헤더만 스타일
            from openpyxl.styles import Alignment, Font, PatternFill
            ws3 = writer.sheets["레짐 분포"]
            fill3 = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            for cell in ws3[1]:
                cell.fill = fill3
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
            from openpyxl.utils import get_column_letter
            for ci, cn in enumerate(df_summary.columns, 1):
                ws3.column_dimensions[get_column_letter(ci)].width = 22

    except ImportError:
        # openpyxl 없으면 스타일 없이 저장
        with pd.ExcelWriter(str(SAVE_PATH)) as writer:
            df_result.to_excel(writer, sheet_name="공통빈도 분석",    index=False)
            df_strong.to_excel(writer, sheet_name="경험적 근거 강함", index=False)
            df_summary.to_excel(writer, sheet_name="레짐 분포",       index=False)

    # ── 저장 요약 ──────────────────────────────────────────────
    grade_counts = df_result["empirical_grade"].value_counts()
    print("\n" + "=" * 60)
    print("  [저장 완료]")
    print("=" * 60)
    print(f"  경로   : {SAVE_PATH}")
    print(f"  총 행수: {len(df_result)}")
    for grade in ["경험적 근거 강함", "경험적 근거 약함", "공통발생 없음"]:
        cnt = grade_counts.get(grade, 0)
        print(f"    {grade:<15}: {cnt}")


# ══════════════════════════════════════════════════════════════
# Step 4: 시나리오 기반 외생변수 생성
# ══════════════════════════════════════════════════════════════

def _make_label_maps(df_regime: pd.DataFrame) -> dict[str, dict[int, str]]:
    """각 변수 HMM 상태 수(K)에 따라 의미적 레이블 매핑 반환.

    scenario_selection.py 와 동일한 규칙 (circular import 방지를 위해 인라인).
    """
    VKOSPI_LABELS = {
        2: {0: "Normal", 1: "Extreme"},
        3: {0: "Low",    1: "Normal",  2: "Extreme"},
    }
    DEFAULT_LABELS = {
        2: {0: "Low", 1: "High"},
        3: {0: "Low", 1: "Mid",  2: "High"},
    }
    maps: dict[str, dict[int, str]] = {}
    for regime_col, display_col in _COL_MAP.items():
        if regime_col not in df_regime.columns:
            continue
        K = int(df_regime[regime_col].dropna().nunique())
        if regime_col == "VKOSPI_resid_regime":
            state_map = VKOSPI_LABELS.get(K, {i: f"S{i}" for i in range(K)})
        else:
            state_map = DEFAULT_LABELS.get(K, {i: f"S{i}" for i in range(K)})
        maps[display_col] = state_map
    return maps


def _build_daily_labeled(
    df_regime: pd.DataFrame,
    label_maps: dict[str, dict[int, str]],
) -> pd.DataFrame:
    """정수 HMM 상태 → 의미적 문자열 레이블로 변환한 일별 DataFrame 반환.

    5개 변수 모두 유효한 행만 유지.
    """
    rows: dict[str, pd.Series] = {}
    for regime_col, display_col in _COL_MAP.items():
        if regime_col not in df_regime.columns:
            continue
        rows[display_col] = (
            df_regime[regime_col]
            .dropna()
            .astype(int)
            .map(label_maps[display_col])
        )
    df = pd.concat(rows, axis=1).dropna()
    df.index.name = "Date"
    return df


def build_scenario_exog(
    df_regime: pd.DataFrame,
    scenario_csv: Path = SCENARIO_CSV_PATH,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """사전 정의된 시나리오 CSV 기반으로 외생변수 더미 DataFrame 생성.

    Parameters
    ----------
    df_regime    : build_regime_df() 반환 DataFrame
    scenario_csv : final_scenarios_latest.csv 경로

    Returns
    -------
    df_exog   : DatetimeIndex × Scenario_ID 바이너리 더미 (0/1)
    df_labeled: DatetimeIndex × 의미적 레이블 (Low/Mid/High 등)
    """
    if not scenario_csv.exists():
        raise FileNotFoundError(
            f"시나리오 CSV 없음: {scenario_csv}\n"
            "analysis/scenario_selection.py 를 먼저 실행하세요."
        )

    scenarios = pd.read_csv(scenario_csv)
    print(f"  시나리오 로드: {len(scenarios)}개  ({scenario_csv.name})")

    label_maps = _make_label_maps(df_regime)
    print("\n  레이블 매핑:")
    for col, smap in label_maps.items():
        print(f"    {col:15s} : {smap}")

    df_labeled = _build_daily_labeled(df_regime, label_maps)
    T = len(df_labeled)
    print(f"\n  완전 관측 일수 (T) : {T}")

    active_cols = [c for c in _DISPLAY_COLS if c in df_labeled.columns]
    exog_cols: dict[str, pd.Series] = {}

    print("\n  시나리오별 활성 일수:")
    for _, row in scenarios.iterrows():
        sid   = str(row["Scenario_ID"])
        mask  = pd.Series(True, index=df_labeled.index)
        for col in active_cols:
            val = row.get(col)
            if pd.notna(val) and str(val) not in ("", "-"):
                mask &= df_labeled[col] == str(val)
        exog_cols[sid] = mask.astype(int)
        n_active   = int(mask.sum())
        combo_label = row.get("combo_label", "")
        print(f"    {sid}: {n_active:4d}일 ({n_active/T:.1%})  [{combo_label}]")

    df_exog = pd.DataFrame(exog_cols, index=df_labeled.index)

    # 어느 시나리오에도 해당하지 않는 날
    n_no_scenario = int((df_exog.sum(axis=1) == 0).sum())
    print(f"\n  미분류(어떤 시나리오에도 해당 없음): {n_no_scenario}일 ({n_no_scenario/T:.1%})")

    return df_exog, df_labeled


def save_exog(
    df_exog: pd.DataFrame,
    df_labeled: pd.DataFrame,
    save_dir: Path = EXOG_SAVE_PATH,
) -> None:
    """외생변수 DataFrame 을 CSV + Excel 2시트로 저장."""
    save_dir.mkdir(parents=True, exist_ok=True)

    csv_path  = save_dir / "scenario_exog_vars.csv"
    xlsx_path = save_dir / "scenario_exog_vars.xlsx"

    df_exog.to_csv(csv_path)
    print(f"\n  CSV 저장 : {csv_path}")

    df_combined = pd.concat([df_labeled, df_exog], axis=1)

    try:
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        with pd.ExcelWriter(str(xlsx_path), engine="openpyxl") as writer:
            df_exog.to_excel(writer,    sheet_name="시나리오 더미")
            df_combined.to_excel(writer, sheet_name="레이블+더미")
            df_labeled.to_excel(writer,  sheet_name="의미적 레이블")

            # 더미 시트: 1인 셀 강조
            ws = writer.sheets["시나리오 더미"]
            fill_on  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            fill_hdr = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            for cell in ws[1]:
                cell.fill = fill_hdr
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if cell.value == 1:
                        cell.fill = fill_on
            for ci in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(ci)].width = 12

    except ImportError:
        with pd.ExcelWriter(str(xlsx_path)) as writer:
            df_exog.to_excel(writer,    sheet_name="시나리오 더미")
            df_combined.to_excel(writer, sheet_name="레이블+더미")
            df_labeled.to_excel(writer,  sheet_name="의미적 레이블")

    print(f"  Excel 저장: {xlsx_path}")


# ══════════════════════════════════════════════════════════════
# 메인 실행
# ══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    # ── HMM 레짐 DataFrame 구축 ────────────────────────────────
    df_regime = build_regime_df(n_init=10, B=1000)

    # ── 공통빈도 지표 계산 ─────────────────────────────────────
    print("\n" + "=" * 60)
    print("  [Step 5] 공통빈도 지표 계산")
    print("=" * 60)
    df_result = compute_pairwise_cooccurrence(df_regime)

    # ── 결과 미리보기 ──────────────────────────────────────────
    print("\n[Lift 상위 15 — 경험적 근거 강함]")
    top = (
        df_result[df_result["empirical_grade"] == "경험적 근거 강함"]
        .nlargest(15, "Lift")
    )
    if len(top) > 0:
        print(top[[
            "var_A", "regime_A", "var_B", "regime_B",
            "n_AB", "T", "P_AB", "Lift", "empirical_grade",
        ]].to_string(index=False))
    else:
        print("  (해당 없음)")

    print("\n[Lift 하위 10 — 경험적 근거 약함 / 공통발생 없음]")
    low = df_result[df_result["empirical_grade"] == "경험적 근거 약함"]
    print(low.head(10)[[
        "var_A", "regime_A", "var_B", "regime_B",
        "n_AB", "T", "Lift", "empirical_grade",
    ]].to_string(index=False))

    # ── Excel 저장 ─────────────────────────────────────────────
    save_to_excel(df_result, df_regime)

    # ── 시나리오 기반 외생변수 생성 ────────────────────────────
    print("\n" + "=" * 60)
    print("  [Step 6] 시나리오 기반 외생변수 생성")
    print("=" * 60)
    if SCENARIO_CSV_PATH.exists():
        df_exog, df_labeled = build_scenario_exog(df_regime)
        save_exog(df_exog, df_labeled)
    else:
        print(f"  [!] 시나리오 CSV 없음: {SCENARIO_CSV_PATH}")
        print("      analysis/scenario_selection.py 를 먼저 실행하세요.")
