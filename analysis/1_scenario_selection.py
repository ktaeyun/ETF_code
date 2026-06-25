"""
analysis/scenario_selection.py
=================================
ETF 시나리오 선택 파이프라인

MODE A: HMM 레짐 분류 → regime_classified_table.csv 생성 (최초 1회)
MODE B: 저장된 테이블 로드 → Step 3~5 시나리오 선택 실행

실행 예시:
  python analysis/scenario_selection.py            # 자동 모드 감지
  python analysis/scenario_selection.py --mode b   # MODE B 강제 (테이블 없으면 에러)
  python analysis/scenario_selection.py --diversity 3 --lift 1.1
"""

from __future__ import annotations

import argparse
import itertools
import sys
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd

_ROOT = Path(__file__).resolve().parent.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from analysis.pairwise_regime_cooccurrence import load_or_build_regime_df


# ══════════════════════════════════════════════════════════════
# 상수
# ══════════════════════════════════════════════════════════════

COL_MAP: dict[str, str] = {
    "Global_RV_regime":        "Bitcoin_RV",
    "VKOSPI_resid_regime":     "VKOSPI",
    "btc_volume_btc_regime":   "KR_Volume",
    "domestic_btc_svi_regime": "KR_SVI",
    "global_btc_svi_regime":   "Global_SVI",
}
DISPLAY_COLS = ["Bitcoin_RV", "VKOSPI", "KR_Volume", "KR_SVI", "Global_SVI"]

# 이론적으로 정의된 전체 레이블 집합 (관측 여부와 무관)
# enumerate_all_theoretical_combos() 에서 사용 — 실제 HMM 상태 수에 맞게 수정 가능
THEORETICAL_STATES: dict[str, list[str]] = {
    "Bitcoin_RV": ["Low", "Mid", "High"],
    "VKOSPI":     ["Normal", "Extreme"],
    "KR_Volume":  ["Low", "High"],
    "KR_SVI":     ["Low", "High"],
    "Global_SVI": ["Low", "Mid", "High"],
}

CLASSIFIED_TABLE_PATH = _ROOT / "results" / "regime_classified_table.csv"
SCENARIO_DIR          = _ROOT / "results" / "scenario_selection"

PAIRWISE_JUDGEMENT_PATH = _ROOT / "results" / "scenario_pair_possible_ambiguity.csv"

# scenario_pair_possible_ambiguity.csv 의 한국어 레이블 → 영어 매핑
_VAR_NAME_MAP: dict[str, str] = {
    "KOSPI_Vol":     "VKOSPI",
    "BTC_RV":        "Bitcoin_RV",
    "KR_BTC_Volume": "KR_Volume",
    "KR_SVI":        "KR_SVI",
    "Global_SVI":    "Global_SVI",
}
_REGIME_LABEL_MAP: dict[str, str] = {
    "극단":     "Extreme",
    "평시":     "Normal",
    "고변동":   "High",
    "중간 변동": "Mid",
    "저변동":   "Low",
    "고유동성": "High",
    "저유동성": "Low",
    "고관심":   "High",
    "저관심":   "Low",
    "보통관심": "Mid",
}


# ══════════════════════════════════════════════════════════════
# Step 1 보조: 레이블 매핑
# ══════════════════════════════════════════════════════════════

def _make_label_maps(df_regime: pd.DataFrame) -> dict[str, dict[int, str]]:
    """각 변수의 고유 상태 수(K)를 기반으로 의미적 레이블 매핑 반환.

    - VKOSPI_resid : K=2 → Normal/Extreme,  K=3 → Low/Normal/Extreme
    - 나머지        : K=2 → Low/High,        K=3 → Low/Mid/High
    """
    VKOSPI_LABELS = {
        2: {0: "Normal", 1: "Extreme"},
        3: {0: "Low", 1: "Normal", 2: "Extreme"},
    }
    DEFAULT_LABELS = {
        2: {0: "Low", 1: "High"},
        3: {0: "Low", 1: "Mid", 2: "High"},
    }

    maps: dict[str, dict[int, str]] = {}
    for regime_col, display_col in COL_MAP.items():
        if regime_col not in df_regime.columns:
            continue
        K = int(df_regime[regime_col].dropna().nunique())
        if regime_col == "VKOSPI_resid_regime":
            state_map = VKOSPI_LABELS.get(K, {i: f"S{i}" for i in range(K)})
        else:
            state_map = DEFAULT_LABELS.get(K, {i: f"S{i}" for i in range(K)})
        maps[display_col] = state_map
    return maps


def _build_daily_labeled(
    df_regime: pd.DataFrame,
    label_maps: dict[str, dict[int, str]],
) -> pd.DataFrame:
    """정수 레짐 레이블 → 의미적 문자열 레이블로 변환한 일별 DataFrame 반환.

    5개 변수 모두 유효한 행만 유지.
    """
    rows: dict[str, pd.Series] = {}
    for regime_col, display_col in COL_MAP.items():
        if regime_col not in df_regime.columns:
            continue
        smap = label_maps[display_col]
        rows[display_col] = (
            df_regime[regime_col]
            .dropna()
            .astype(int)
            .map(smap)
        )
    df = pd.concat(rows, axis=1).dropna()
    df.index.name = "Date"
    return df


# ══════════════════════════════════════════════════════════════
# 쌍별 CSV 판정 로드 및 5변수 조합 재분류
# ══════════════════════════════════════════════════════════════

def _build_pairwise_lookup(csv_path: Path) -> dict[tuple, str]:
    """scenario_pair_possible_ambiguity.csv → (var_A, r_A, var_B, r_B): judgement 사전.

    한국어 레이블을 영어로 번역하고, 방향 무관하게 (A,B), (B,A) 양방향 등록.
    """
    df = pd.read_csv(csv_path)
    lookup: dict[tuple, str] = {}
    for _, row in df.iterrows():
        var_r  = _VAR_NAME_MAP.get(row["row_variable"],    row["row_variable"])
        var_c  = _VAR_NAME_MAP.get(row["column_variable"], row["column_variable"])
        reg_r  = _REGIME_LABEL_MAP.get(row["row_regime"],    row["row_regime"])
        reg_c  = _REGIME_LABEL_MAP.get(row["column_regime"], row["column_regime"])
        j      = row["judgement"]
        lookup[(var_r, reg_r, var_c, reg_c)] = j
        lookup[(var_c, reg_c, var_r, reg_r)] = j   # 역방향도 동일
    return lookup


def reclassify_from_pairwise_csv(
    df_combo:  pd.DataFrame,
    csv_path:  Path = PAIRWISE_JUDGEMENT_PATH,
) -> pd.DataFrame:
    """5변수 조합 테이블의 theoretical_class를 쌍별 CSV 판정으로 재계산.

    10개 쌍 중 하나라도 Ambiguity → 해당 조합 Ambiguity.
    모두 Possible → Possible.
    쌍이 CSV에 없으면 'Unknown' (해당 조합 Ambiguity 처리).
    """
    lookup = _build_pairwise_lookup(csv_path)

    def _classify(row: pd.Series) -> str:
        for col_i, col_j in itertools.combinations(DISPLAY_COLS, 2):
            key = (col_i, row[col_i], col_j, row[col_j])
            j   = lookup.get(key, "Ambiguity")   # 미등재 = Ambiguity
            if j != "Possible":
                return "Ambiguity"
        return "Possible"

    df = df_combo.copy()
    df["theoretical_class"] = df.apply(_classify, axis=1)

    n_p = (df["theoretical_class"] == "Possible").sum()
    n_a = (df["theoretical_class"] == "Ambiguity").sum()
    print(f"  쌍별 CSV 재분류 완료 → Possible: {n_p}개  Ambiguity: {n_a}개")
    return df


# ══════════════════════════════════════════════════════════════
# Step 2 보조: 공통빈도 지표
# ══════════════════════════════════════════════════════════════

def _precompute_pairwise_lifts(df_labeled: pd.DataFrame) -> dict[tuple, float]:
    """모든 변수쌍 × 레이블쌍 lift 사전 계산."""
    T = len(df_labeled)
    lift_dict: dict[tuple, float] = {}
    for col_i, col_j in itertools.combinations(DISPLAY_COLS, 2):
        if col_i not in df_labeled or col_j not in df_labeled:
            continue
        for r_i in df_labeled[col_i].unique():
            for r_j in df_labeled[col_j].unique():
                n_i  = (df_labeled[col_i] == r_i).sum()
                n_j  = (df_labeled[col_j] == r_j).sum()
                n_ij = ((df_labeled[col_i] == r_i) & (df_labeled[col_j] == r_j)).sum()
                if n_i > 0 and n_j > 0:
                    lift = (n_ij / T) / ((n_i / T) * (n_j / T))
                else:
                    lift = np.nan
                lift_dict[(col_i, r_i, col_j, r_j)] = lift
    return lift_dict


def _pairwise_lift_mean(combo: dict[str, str], lift_dict: dict[tuple, float]) -> float:
    """하나의 조합에 대해 모든 변수쌍 lift의 평균값 반환."""
    lifts: list[float] = []
    for col_i, col_j in itertools.combinations(DISPLAY_COLS, 2):
        if col_i not in combo or col_j not in combo:
            continue
        key = (col_i, combo[col_i], col_j, combo[col_j])
        val = lift_dict.get(key, np.nan)
        lifts.append(val)
    return float(np.nanmean(lifts)) if lifts else np.nan


# ══════════════════════════════════════════════════════════════
# MODE A: 분류 테이블 구축
# ══════════════════════════════════════════════════════════════

def build_classified_table(n_init: int = 10, B: int = 1000) -> pd.DataFrame:
    """HMM 레짐 분류 → regime_classified_table.csv 생성.

    Returns
    -------
    pd.DataFrame : 조합별 1행, 컬럼 = DISPLAY_COLS + 통계 컬럼
    """
    print("=" * 60)
    print("  [MODE A] 레짐 분류 테이블 구축")
    print("=" * 60)

    df_regime = load_or_build_regime_df(n_init=n_init, B=B)
    label_maps = _make_label_maps(df_regime)

    print("\n  레이블 매핑:")
    for col, smap in label_maps.items():
        print(f"    {col:15s} : {smap}")

    df_labeled = _build_daily_labeled(df_regime, label_maps)
    T = len(df_labeled)
    print(f"\n  완전 관측 일수 (T) : {T}")

    lift_dict = _precompute_pairwise_lifts(df_labeled)

    # 조합별 집계
    grouped = (
        df_labeled.groupby(DISPLAY_COLS, observed=True)
        .size()
        .reset_index(name="joint_frequency")
    )
    grouped["joint_prob"] = grouped["joint_frequency"] / T
    grouped["combo_label"] = grouped[DISPLAY_COLS].agg("-".join, axis=1)

    # pairwise_lift_mean 계산
    def _row_lift_mean(row: pd.Series) -> float:
        combo = {c: row[c] for c in DISPLAY_COLS}
        return _pairwise_lift_mean(combo, lift_dict)

    grouped["pairwise_lift_mean"] = grouped.apply(_row_lift_mean, axis=1)

    # Possible/Ambiguity 판정: scenario_pair_possible_ambiguity.csv 기반
    if PAIRWISE_JUDGEMENT_PATH.exists():
        grouped = reclassify_from_pairwise_csv(grouped, PAIRWISE_JUDGEMENT_PATH)
    else:
        print(f"  [경고] {PAIRWISE_JUDGEMENT_PATH.name} 없음 → lift 기반으로 대체")
        grouped["theoretical_class"] = np.where(
            grouped["pairwise_lift_mean"] >= 1.0, "Possible", "Ambiguity"
        )

    col_order = ["combo_label"] + DISPLAY_COLS + [
        "joint_frequency", "joint_prob", "pairwise_lift_mean", "theoretical_class"
    ]
    df_result = grouped[col_order].sort_values(
        "joint_frequency", ascending=False
    ).reset_index(drop=True)

    CLASSIFIED_TABLE_PATH.parent.mkdir(parents=True, exist_ok=True)
    df_result.to_csv(CLASSIFIED_TABLE_PATH, index=False)

    n_possible   = (df_result["theoretical_class"] == "Possible").sum()
    n_ambiguity  = (df_result["theoretical_class"] == "Ambiguity").sum()
    print(f"\n  총 조합 수     : {len(df_result)}")
    print(f"  Possible       : {n_possible}")
    print(f"  Ambiguity      : {n_ambiguity}")
    print(f"  저장 경로      : {CLASSIFIED_TABLE_PATH}")

    return df_result


# ══════════════════════════════════════════════════════════════
# 테이블 로드
# ══════════════════════════════════════════════════════════════

def load_classified_table() -> pd.DataFrame:
    """저장된 regime_classified_table.csv 로드."""
    if not CLASSIFIED_TABLE_PATH.exists():
        raise FileNotFoundError(
            f"분류 테이블이 없습니다: {CLASSIFIED_TABLE_PATH}\n"
            "MODE A를 먼저 실행하세요: python analysis/scenario_selection.py --mode a"
        )
    df = pd.read_csv(CLASSIFIED_TABLE_PATH)
    print(f"  분류 테이블 로드: {len(df)}개 조합  ({CLASSIFIED_TABLE_PATH})")
    return df


# ══════════════════════════════════════════════════════════════
# Step 3: 핵심 시나리오 선택
# ══════════════════════════════════════════════════════════════

def _count_shared_states(a: pd.Series, b: pd.Series) -> int:
    """두 조합 행(Series) 간 동일 레이블 변수 수 반환."""
    return sum(a[col] == b[col] for col in DISPLAY_COLS)


def step3_core(
    df_combo:             pd.DataFrame,
    diversity_constraint: int | None = None,
    lift_threshold:       float      = 1.0,
    freq_threshold:       int        = 1,
) -> pd.DataFrame:
    """Step 3: CSV 판정 기준 Possible 조합 전체 반환.

    diversity_constraint=None(기본값): 제약 없이 Possible 전체 반환.
    """
    print("\n" + "=" * 60)
    print("  [Step 3] 핵심 시나리오 선택")
    constraint_str = "없음(전체)" if diversity_constraint is None else str(diversity_constraint)
    print(f"  diversity_constraint={constraint_str}  "
          f"lift_threshold={lift_threshold}  freq_threshold={freq_threshold}")
    print("=" * 60)

    # 입력 필터: CSV 판정 Possible + 최소 빈도 (lift는 참조 지표, 필터 제외)
    candidates = df_combo[
        (df_combo["theoretical_class"] == "Possible") &
        (df_combo["joint_frequency"] >= freq_threshold)
    ].sort_values("joint_frequency", ascending=False).reset_index(drop=True)

    print(f"  후보 수 (Possible, freq ≥ {freq_threshold}): "
          f"{len(candidates)}")

    if diversity_constraint is None:
        # 제약 없이 전체 선택
        selected = [row for _, row in candidates.iterrows()]
        skipped  = 0
    else:
        selected: list[pd.Series] = []
        skipped = 0
        for _, row in candidates.iterrows():
            if all(
                _count_shared_states(row, sel) <= diversity_constraint
                for sel in selected
            ):
                selected.append(row)
            else:
                skipped += 1

    if not selected:
        print("  → 선택된 핵심 시나리오 없음")
        return pd.DataFrame(columns=["Rank"] + DISPLAY_COLS + [
            "joint_frequency", "pairwise_lift_mean", "theoretical_class", "combo_label"
        ])

    df_core = pd.DataFrame(selected).reset_index(drop=True)
    df_core.insert(0, "Rank", range(1, len(df_core) + 1))

    skipped_msg = f"  |  다양성 제약으로 제외: {skipped}개" if diversity_constraint is not None else ""
    print(f"  선택: {len(df_core)}개{skipped_msg}")

    _print_table(df_core, ["Rank"] + DISPLAY_COLS + [
        "joint_frequency", "pairwise_lift_mean", "theoretical_class"
    ])
    return df_core


# ══════════════════════════════════════════════════════════════
# Step 4: 최종 요약
# ══════════════════════════════════════════════════════════════

def step5_validate(
    core_df:   pd.DataFrame,
    prev_path: Path | None = None,
) -> pd.DataFrame:
    """Step 4(구 Step 5): 핵심 시나리오 최종 요약 테이블 생성."""
    print("\n" + "=" * 60)
    print("  [Step 4] 최종 요약")
    print("=" * 60)

    final_rows: list[dict] = []
    for _, r in core_df.iterrows():
        final_rows.append({
            "Scenario_ID": f"S{int(r['Rank']):02d}",
            **{c: r[c] for c in DISPLAY_COLS},
            "joint_frequency":    int(r["joint_frequency"]),
            "pairwise_lift_mean": round(float(r["pairwise_lift_mean"]), 4),
            "theoretical_class":  r["theoretical_class"],
            "combo_label":        r["combo_label"],
        })

    final_df = pd.DataFrame(final_rows)

    print(f"\n  선택된 시나리오: {len(final_df)}개")
    _print_table(final_df, ["Scenario_ID"] + DISPLAY_COLS + [
        "joint_frequency", "pairwise_lift_mean", "theoretical_class"
    ])

    if prev_path and prev_path.exists():
        _print_diff(final_df, prev_path)

    return final_df


def _print_diff(current_df: pd.DataFrame, prev_path: Path) -> None:
    """이전 시나리오 결과와 비교 Diff 출력."""
    prev_df = pd.read_csv(prev_path)
    print("\n  [Diff — 이전 vs 현재]")

    prev_map = prev_df.set_index("Scenario_ID")["combo_label"].to_dict()
    curr_map = current_df.set_index("Scenario_ID")["combo_label"].to_dict()

    all_ids = sorted(set(prev_map) | set(curr_map))
    diff_rows = []
    for sid in all_ids:
        prev_val = prev_map.get(sid, "(없음)")
        curr_val = curr_map.get(sid, "(없음)")
        if prev_val != curr_val:
            change = "변경" if prev_val != "(없음)" and curr_val != "(없음)" else (
                "추가" if prev_val == "(없음)" else "제거"
            )
            diff_rows.append({
                "Scenario_ID": sid, "Previous": prev_val,
                "Updated": curr_val, "Change": change
            })

    if diff_rows:
        print(pd.DataFrame(diff_rows).to_string(index=False))
    else:
        print("  변경 없음 (이전과 동일)")


def _print_table(df: pd.DataFrame, cols: list[str]) -> None:
    display_cols = [c for c in cols if c in df.columns]
    print(df[display_cols].to_string(index=False))


# ══════════════════════════════════════════════════════════════
# Excel 저장
# ══════════════════════════════════════════════════════════════

_CLASS_COLORS = {"Possible": "C6EFCE", "Ambiguity": "FFC7CE"}


def _style_ws(ws, df: pd.DataFrame, color_col: str, color_map: dict) -> None:
    try:
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return

    hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    for ci, cn in enumerate(df.columns, 1):
        max_w = max(len(str(cn)),
                    df[cn].astype(str).str.len().max() if len(df) > 0 else 0)
        ws.column_dimensions[get_column_letter(ci)].width = min(max_w + 2, 30)

    if color_col not in df.columns:
        return
    coli = df.columns.get_loc(color_col) + 1
    for ri in range(2, len(df) + 2):
        val   = ws.cell(row=ri, column=coli).value
        color = color_map.get(val)
        if color:
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            for ci in range(1, len(df.columns) + 1):
                ws.cell(row=ri, column=ci).fill = fill


def save_to_excel(final_df: pd.DataFrame) -> Path:
    """시나리오 결과를 Excel로 저장."""
    SCENARIO_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path  = SCENARIO_DIR / f"scenario_selection_{timestamp}.xlsx"

    prev_csv = SCENARIO_DIR / "final_scenarios_latest.csv"
    if prev_csv.exists():
        prev_csv.rename(SCENARIO_DIR / "final_scenarios_prev.csv")
    final_df.to_csv(prev_csv, index=False)

    try:
        with pd.ExcelWriter(str(out_path), engine="openpyxl") as writer:
            final_df.to_excel(writer, sheet_name="Scenarios", index=False)
            _style_ws(writer.sheets["Scenarios"], final_df,
                      "theoretical_class", _CLASS_COLORS)
    except ImportError:
        with pd.ExcelWriter(str(out_path)) as writer:
            final_df.to_excel(writer, sheet_name="Scenarios", index=False)

    print(f"\n  Excel 저장: {out_path}")
    return out_path


# ══════════════════════════════════════════════════════════════
# MODE B 실행 엔트리포인트
# ══════════════════════════════════════════════════════════════

def run_mode_b(
    diversity_constraint: int | None = None,
    lift_threshold:       float      = 1.0,
    freq_threshold:       int        = 1,
    include_unobserved:   bool       = False,
) -> dict:
    """저장된 분류 테이블을 기반으로 Step 3~5 실행.

    include_unobserved=True: 이론적으로 Possible이지만 관측 빈도=0인 조합도 포함.
    """
    print("=" * 60)
    print("  [MODE B] 시나리오 선택 (저장된 테이블 사용)")
    if include_unobserved:
        print("  (미관측 Possible 조합 포함)")
    print("=" * 60)

    df_combo = load_classified_table()

    if include_unobserved and PAIRWISE_JUDGEMENT_PATH.exists():
        # 이론적 전체 조합 생성 (관측+미관측) 후 Possible 필터
        df_source = enumerate_all_theoretical_combos(df_combo)
        used_freq_threshold = 0  # 미관측(freq=0) 포함
    else:
        df_source = df_combo
        used_freq_threshold = freq_threshold

    core_df = step3_core(
        df_source,
        diversity_constraint=diversity_constraint,
        lift_threshold=lift_threshold,
        freq_threshold=used_freq_threshold,
    )

    prev_csv = SCENARIO_DIR / "final_scenarios_prev.csv"
    final_df = step5_validate(
        core_df,
        prev_path=prev_csv if prev_csv.exists() else None,
    )

    out_path = save_to_excel(final_df)

    return {
        "core":  core_df,
        "final": final_df,
        "path":  out_path,
    }


# ══════════════════════════════════════════════════════════════
# 전체 이론적 조합 열거 (AXES 없이 모두 Possible 검토)
# ══════════════════════════════════════════════════════════════

def enumerate_all_theoretical_combos(
    df_combo:  pd.DataFrame,
    csv_path:  Path = PAIRWISE_JUDGEMENT_PATH,
) -> pd.DataFrame:
    """THEORETICAL_STATES 기반 카테시안 곱으로 이론적 전체 조합을 생성.

    분류 기준: scenario_pair_possible_ambiguity.csv 쌍별 판정
      - 10쌍 모두 Possible → Possible
      - 하나라도 Ambiguity → Ambiguity

    추가 지표: 관측된 조합에 한해 pairwise_lift_mean 병합
      - 미관측 조합(joint_frequency=0)은 lift=NaN, Possible/Ambiguity 분류는 유지
    """
    observed_states: dict[str, list[str]] = {
        col: sorted(df_combo[col].dropna().unique().tolist())
        for col in DISPLAY_COLS if col in df_combo.columns
    }
    state_sets: dict[str, list[str]] = {
        col: THEORETICAL_STATES.get(col, observed_states.get(col, []))
        for col in DISPLAY_COLS
    }

    print("\n  [변수별 레이블 집합 (THEORETICAL_STATES 기준)]")
    total_combos = 1
    for col, states in state_sets.items():
        obs = observed_states.get(col, [])
        missing = sorted(set(states) - set(obs))
        note = f"  <- 미관측: {missing}" if missing else ""
        print(f"    {col:15s}: {states}{note}")
        total_combos *= len(states)
    print(f"\n  이론적 전체 조합 수: {total_combos}")

    # 카테시안 곱
    cols_ordered = [c for c in DISPLAY_COLS if c in state_sets]
    all_combos = list(itertools.product(*[state_sets[c] for c in cols_ordered]))
    df_all = pd.DataFrame(all_combos, columns=cols_ordered)
    df_all["combo_label"] = df_all[cols_ordered].agg("-".join, axis=1)

    # CSV 쌍별 판정으로 Possible / Ambiguity 분류
    lookup = _build_pairwise_lookup(csv_path)

    def _classify_row(row: pd.Series) -> str:
        for col_i, col_j in itertools.combinations(cols_ordered, 2):
            key = (col_i, row[col_i], col_j, row[col_j])
            if lookup.get(key, "Ambiguity") != "Possible":
                return "Ambiguity"
        return "Possible"

    df_all["theoretical_class"] = df_all.apply(_classify_row, axis=1)

    # 관측 빈도 + lift 병합 (미관측은 freq=0, lift=NaN)
    merge_cols = ["combo_label", "joint_frequency", "joint_prob", "pairwise_lift_mean"]
    obs_cols   = [c for c in merge_cols if c in df_combo.columns]
    if obs_cols:
        df_all = df_all.merge(df_combo[obs_cols], on="combo_label", how="left")
        df_all["joint_frequency"] = df_all["joint_frequency"].fillna(0).astype(int)
        df_all["joint_prob"]      = df_all["joint_prob"].fillna(0.0)
        # pairwise_lift_mean: 미관측은 NaN 유지

    # 정렬: Possible 먼저, 관측 빈도 내림차순
    df_all = df_all.sort_values(
        ["theoretical_class", "joint_frequency"],
        ascending=[True, False],
    ).reset_index(drop=True)
    df_all.insert(0, "No", range(1, len(df_all) + 1))

    n_possible  = (df_all["theoretical_class"] == "Possible").sum()
    n_ambiguity = (df_all["theoretical_class"] == "Ambiguity").sum()
    n_observed  = (df_all["joint_frequency"] > 0).sum()
    n_unobs_pos = ((df_all["theoretical_class"] == "Possible") & (df_all["joint_frequency"] == 0)).sum()
    print(f"  분류 완료 (CSV 쌍별 판정 기준)")
    print(f"    Possible  : {n_possible}개  (이중 미관측: {n_unobs_pos}개, lift=NaN)")
    print(f"    Ambiguity : {n_ambiguity}개")
    print(f"    전체 관측된 조합: {n_observed}개  |  미관측: {total_combos - n_observed}개")

    return df_all


def run_all_combos(freq_threshold: int = 0) -> dict:
    """AXES 사전 정의 없이 이론적으로 가능한 모든 조합을 열거하고
    CSV 쌍별 판정으로 Possible / Ambiguity 분류 후 출력·저장.

    Possible 조합에는 데이터 기반 pairwise_lift_mean 추가 표시.
    미관측 Possible(lift=NaN)은 Ambiguity로 강등하지 않고 Possible 유지.
    'Korean_Specificity' 빈 컬럼 포함 Excel 저장.
    """
    print("=" * 60)
    print("  [ALL COMBOS] 이론적 전체 조합 열거 (CSV 판정 + Lift 참조)")
    print("=" * 60)

    df_combo = load_classified_table()
    df_all   = enumerate_all_theoretical_combos(df_combo)

    if freq_threshold > 0:
        df_all = df_all[df_all["joint_frequency"] >= freq_threshold].reset_index(drop=True)
        df_all["No"] = range(1, len(df_all) + 1)
        print(f"  freq_threshold={freq_threshold} 적용 후: {len(df_all)}개")

    possible_df  = df_all[df_all["theoretical_class"] == "Possible"]
    ambiguity_df = df_all[df_all["theoretical_class"] == "Ambiguity"]

    show_cols = ["No"] + DISPLAY_COLS + [
        "joint_frequency", "joint_prob", "pairwise_lift_mean", "theoretical_class"
    ]
    show_cols = [c for c in show_cols if c in df_all.columns]

    print("\n" + "=" * 60)
    print(f"  [Possible - {len(possible_df)}개]  (CSV 판정 기준, lift는 참조 지표)")
    print("=" * 60)
    _print_table(possible_df, show_cols)

    print("\n" + "=" * 60)
    print(f"  [Ambiguity - {len(ambiguity_df)}개]")
    print("=" * 60)
    _print_table(ambiguity_df, show_cols)

    df_all["Korean_Specificity"] = ""

    SCENARIO_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path  = SCENARIO_DIR / f"all_theoretical_combos_{timestamp}.xlsx"

    save_cols = ["No"] + DISPLAY_COLS + [
        "joint_frequency", "joint_prob", "pairwise_lift_mean",
        "theoretical_class", "Korean_Specificity", "combo_label",
    ]
    save_cols = [c for c in save_cols if c in df_all.columns]

    _class_colors = {"Possible": "C6EFCE", "Ambiguity": "FFC7CE"}

    sheets = [
        ("전체 조합",  df_all),
        ("Possible",   possible_df),
        ("Ambiguity",  ambiguity_df),
    ]

    try:
        with pd.ExcelWriter(str(out_path), engine="openpyxl") as writer:
            for sheet_name, df_s in sheets:
                df_s[save_cols].to_excel(writer, sheet_name=sheet_name, index=False)
            for sheet_name, df_s in sheets:
                _style_ws(writer.sheets[sheet_name], df_s[save_cols],
                          "theoretical_class", _class_colors)
    except ImportError:
        with pd.ExcelWriter(str(out_path)) as writer:
            for sheet_name, df_s in sheets:
                df_s[save_cols].to_excel(writer, sheet_name=sheet_name, index=False)

    print(f"\n  Excel 저장: {out_path}")
    print(f"  -> 'Korean_Specificity' 컬럼에 직접 표시 후 저장하세요.")

    return {
        "all":       df_all,
        "possible":  possible_df,
        "ambiguity": ambiguity_df,
        "path":      out_path,
    }


# ══════════════════════════════════════════════════════════════
# 쌍별 CSV 기반 실행 엔트리포인트
# ══════════════════════════════════════════════════════════════

def run_from_pairs(freq_threshold: int = 1) -> dict:
    """scenario_pair_possible_ambiguity.csv 판정으로 재분류 후 전체 Possible 시나리오 선택.

    - freq_threshold=0 : 미관측(joint_frequency=0) Possible 조합 포함
    - freq_threshold=1 : 실제 관측된 Possible 조합만 (기본값)
    - 한국 특수성 축 매칭 결과를 각 시나리오에 직접 태깅
    """
    print("=" * 60)
    print("  [쌍별 CSV 기반] Possible 시나리오 최대화 선택")
    freq_label = "관측 + 미관측 포함" if freq_threshold == 0 else f"freq ≥ {freq_threshold}"
    print(f"  ({freq_label})")
    print("=" * 60)

    # 1. 테이블 로드
    df_combo = load_classified_table()

    # 2. 이론적 전체 조합 열거 (미관측 freq=0 포함) + 쌍별 CSV 판정
    #    enumerate_all_theoretical_combos 내부에서 pairwise CSV 재분류 수행
    df_all = enumerate_all_theoretical_combos(df_combo)

    # 3. Possible 조합만 추출, freq 기준 적용
    core_df = df_all[
        (df_all["theoretical_class"] == "Possible") &
        (df_all["joint_frequency"] >= freq_threshold)
    ].copy().reset_index(drop=True)

    # Rank 컬럼 삽입 (빈도 내림차순으로 이미 정렬됨)
    if "Rank" not in core_df.columns:
        core_df.insert(0, "Rank", range(1, len(core_df) + 1))
    if "No" in core_df.columns:
        core_df = core_df.drop(columns=["No"])

    n_obs   = (core_df["joint_frequency"] > 0).sum()
    n_unobs = (core_df["joint_frequency"] == 0).sum()
    print(f"\n  선택된 Possible 시나리오: {len(core_df)}개  "
          f"(관측: {n_obs}개 | 미관측: {n_unobs}개)")

    _print_table(core_df, ["Rank"] + DISPLAY_COLS + [
        "joint_frequency", "pairwise_lift_mean", "theoretical_class"
    ])

    # 5. Excel 저장
    SCENARIO_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path  = SCENARIO_DIR / f"scenario_from_pairs_{timestamp}.xlsx"

    save_cols = ["Rank"] + DISPLAY_COLS + [
        "joint_frequency", "joint_prob", "pairwise_lift_mean",
        "theoretical_class", "combo_label",
    ]
    save_cols = [c for c in save_cols if c in core_df.columns]

    try:
        with pd.ExcelWriter(str(out_path), engine="openpyxl") as writer:
            core_df[save_cols].to_excel(writer, sheet_name="Possible 전체", index=False)
            _style_ws(writer.sheets["Possible 전체"], core_df[save_cols],
                      "theoretical_class", _CLASS_COLORS)
    except ImportError:
        with pd.ExcelWriter(str(out_path)) as writer:
            core_df[save_cols].to_excel(writer, sheet_name="Possible 전체", index=False)

    print(f"\n  Excel 저장: {out_path}")
    print(f"  전체 Possible: {len(core_df)}개")

    # 6. final_scenarios_latest.csv 저장 (시뮬레이터 호환 형식)
    final_df = core_df.copy()
    final_df.insert(0, "Scenario_ID", ["P" + str(r).zfill(2) for r in final_df["Rank"]])
    csv_cols = ["Scenario_ID"] + DISPLAY_COLS + [
        "joint_frequency", "pairwise_lift_mean", "theoretical_class", "combo_label",
    ]
    csv_cols = [c for c in csv_cols if c in final_df.columns]

    prev_csv = SCENARIO_DIR / "final_scenarios_latest.csv"
    if prev_csv.exists():
        prev_csv.rename(SCENARIO_DIR / "final_scenarios_prev.csv")
    final_df[csv_cols].to_csv(prev_csv, index=False)
    print(f"  CSV 저장 (시뮬레이터용): {prev_csv}")

    return {
        "possible_all": core_df,
        "path":          out_path,
    }


# ══════════════════════════════════════════════════════════════
# 메인 실행
# ══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="ETF 시나리오 선택 파이프라인")
    parser.add_argument(
        "--mode", choices=["a", "b", "pairs", "all"], default=None,
        help=(
            "a: 테이블 구축, "
            "b: lift 기반 선택, "
            "pairs: 쌍별 CSV 기반 Possible 최대화, "
            "all: 이론적 전체 조합 열거 (AXES 무정의, 사용자 직접 선별용). "
            "미지정 시 자동 감지."
        )
    )
    parser.add_argument("--diversity", type=int,   default=None, metavar="N",
                        help="다양성 제약 (기본값=2)")
    parser.add_argument("--lift",      type=float, default=1.0, metavar="F",
                        help="Lift 임계값 (기본값=1.0)")
    parser.add_argument("--freq",      type=int,   default=1,   metavar="N",
                        help="최소 공통 발생 횟수 (기본값=1)")
    parser.add_argument("--n_init",    type=int,   default=10,
                        help="HMM 초기값 반복 (MODE A 전용, 기본값=10)")
    parser.add_argument("--B",         type=int,   default=1000,
                        help="Bootstrap 반복 (MODE A 전용, 기본값=1000)")
    args = parser.parse_args()
    _include_unobserved = False  # 기본값 (--mode 직접 지정 시)

    # --mode 미지정 시 인터랙티브 메뉴
    if args.mode is None:
        table_exists = CLASSIFIED_TABLE_PATH.exists()
        print("\n" + "=" * 60)
        print("  ETF 시나리오 선택 파이프라인")
        print("=" * 60)
        print(f"  [분류 테이블] {'있음 ✓' if table_exists else '없음 → MODE A 필요'}")
        print()
        print("  [1] MODE A  — HMM 레짐 분류 테이블 새로 구축")
        print("  [2] MODE B  — 저장된 테이블로 Possible 시나리오 선택")
        print("  [3] pairs   — 쌍별 CSV 판정으로 Possible 전수 선택")
        print("  [4] all     — 이론적 전체 조합 열거 (관측 여부 무관)")
        print()
        _CHOICE_MAP = {"1": "a", "2": "b", "3": "pairs", "4": "all",
                       "a": "a", "b": "b", "pairs": "pairs", "all": "all"}
        while True:
            choice = input("  선택 (1~4 또는 a/b/pairs/all): ").strip().lower()
            if choice in _CHOICE_MAP:
                args.mode = _CHOICE_MAP[choice]
                break
            print("  잘못된 입력입니다. 다시 입력해주세요.")

        # MODE B일 때 미관측 포함 여부 추가 질문
        _include_unobserved = False
        if args.mode in ("b", "a"):
            print()
            ans = input("  이론적 Possible이지만 관측 빈도=0인 조합도 포함할까요? (y/N): ").strip().lower()
            _include_unobserved = ans in ("y", "yes")
        print()

    if args.mode == "a":
        build_classified_table(n_init=args.n_init, B=args.B)
        run_mode_b(
            diversity_constraint=args.diversity,
            lift_threshold=args.lift,
            freq_threshold=args.freq,
            include_unobserved=_include_unobserved,
        )
    elif args.mode == "pairs":
        run_from_pairs(freq_threshold=args.freq)
    elif args.mode == "all":
        freq_all = args.freq if args.freq != 1 else 0
        run_all_combos(freq_threshold=freq_all)
    else:
        run_mode_b(
            diversity_constraint=args.diversity,
            lift_threshold=args.lift,
            freq_threshold=args.freq,
            include_unobserved=_include_unobserved,
        )
